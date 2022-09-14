"""
Create a set of JSON files from a set of XLSX files for the M-layer register.

The process will require iterations.

    * Copy the JSON folder under the builder directory.

    * Run `to_xlsx.py` to create a set of XLSX files that reflect the current state of the JSON files.

    * Open all of the XLSX files at the same time and allow them to update content. Then save them again.

    * Introduce new rows and/or sheets to the various workbooks to define objects or transformations. 
    
        NOTE: 
        When introducing new objects that need UUIDs, enter the word None in place of the UUID. 
        This script will automatically generate the missing UUIDs.
        
        It is not possible to use cell references until UUIDs have been assigned. So, new objects 
        requiring UUIDS should be created by running this script. Then followed by `to_xlsx.py` to 
        create an updated set of XLSX files. 

        Only one layer of missing UUIDs can be handled at a time. So, do references first, then update
        and start again, then do the scales, update and start again. Finally, do the conversions.
    
        When setting up conversions, or other parts of the register that refer to unique objects, 
        use Excel cell references (type '=' in the cell and then click on the target cell).
        
        To update all the workbooks, open them all then from the task bar Excel grouped button 
        right-click and select close all. Then accept all requests to save the workbooks.
    
    * When finished, copy the JSON folder, under the builder directory, to replace the JSON folder under the m_layer directory.  

"""
import json 
import glob
import os
import uuid
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter 

from ast import literal_eval
     
# --------------------------------------------------------------------------- 
root = os.path.abspath( os.path.dirname(__file__) )

json_root = os.path.join( root, r'json' )
xl_root = os.path.join( root, 'xl' )

# --------------------------------------------------------------------------- 
def xl_path(filename):
    """
    Return the path to an XLSX file
    """
    return os.path.abspath( 
        os.path.join(
            xl_root,
            '{}.xlsx'.format(filename)
        ) 
    )
           
# --------------------------------------------------------------------------- 
def get_uid(file,ws,r_value,i,col):
    """
    `r_value` should be an M-layer UID
    If the UUID element has not been assigned it is generated here
    
    """
    uid = literal_eval( r_value )
    
    # A new UUID will be assigned if necessary
    if uid[1] is None:
        uid[1] = uuid.uuid4().int
        ws.cell(row=i,column=col,value=str(uid))        
 
    return uid  
      
# --------------------------------------------------------------------------- 
def put_uid(file,ws,r_value,i,col):
    """
    `r_value` is expected to be a fully formed M-layer uid.
    If it is not this will fail.
    
    """
    if r_value is None:
        col_letter = get_column_letter(col)
        addr = "{}{}".format(col_letter,i)
        print("no uid: [{}.xlsx]{}!{}".format(file,ws.title,addr))
        assert False
    else:
        if r_value[1] is None:
            col_letter = get_column_letter(col)
            addr = "{}{}".format(col_letter,i)
            print( "incomplete uid: [{}.xlsx]{}!{}".format(file,ws.title,addr) )
            assert False
        else:
            return literal_eval( r_value )            
            
# --------------------------------------------------------------------------- 
def dump_to_json(directory,filename,lst):
    """
    Write `lst` to a JSON file 
    
    """
    dir = os.path.abspath(
        os.path.join(   
            json_root,
            directory
        )
    )
    if not os.path.isdir(dir):
        os.mkdir(dir)
    
    path = os.path.abspath(
        os.path.join( 
            dir,
            '{}.json'.format(filename)
        )
    ) 
    
    with open( path, 'w') as f:
        json.dump(lst,f,indent=4)
 
# =========================================================================== 
if __name__ == '__main__': 

    root = os.path.abspath( os.path.dirname(__file__) )

    if not os.path.isdir( json_root ):
        os.mkdir( json_root )
    
    # ---------------------------------------------------------------------------
    # Process aspects 
    # 
    # obj = {
        # '__entry__' : 'Aspect',        
        # 'uid' : [], 
        # 'locale' : dict( default =  dict( name="", symbol = "" )) , 
        # 'metadata' : dict( url = "" )
    # }
    
    directory = "aspects"   
    def to_aspect(ws,r,i):
 
        uid = get_uid(directory,ws,r[0].value,i,0+1)
 
        obj = dict(
            __entry__ = 'Aspect',
            uid = uid,
            locale = dict( 
                default = dict(
                    name = str() if r[1].value is None else str( r[1].value ),
                    symbol = str() if r[1].value is None else str( r[2].value )
                )
            )
        )
        
        # metadata may not be provided
        if r[3].value is not None:
            obj['metadata'] = dict( url = str( r[3].value ) )
            
        return obj
        
    wb = load_workbook( xl_path(directory), data_only=True )
    for s in wb.sheetnames:
        lst = []
        ws = wb[s]
        for i,r in enumerate(ws.rows):
            if i == 0: 
                continue
            else:
                lst.append( to_aspect(ws,r,i+1) )
            
            dump_to_json(directory,s,lst)

    wb.save( xl_path(directory) )   # There may have been UUIDs added
            
    # ---------------------------------------------------------------------------
    # Process systems 
    # 
    # obj = {
        # '__entry__' : 'UnitSystem',        
        # 'uid' : [], 
        # 'name' : "", 
        # 'basis' : []
    # }
    
    directory = "systems"
    def to_system(ws,r,i):

        uid = get_uid(directory,ws,r[0].value,i,0+1)

        return dict(
            __entry__ = 'UnitSystem',
            uid = uid,
            name = r[1].value,
            basis = literal_eval( r[2].value )
        )
        
    wb = load_workbook( xl_path(directory), data_only=True  )
    for s in wb.sheetnames:
        lst = []
        ws = wb[s]
        for i,r in enumerate(ws.rows):
            if i == 0: 
                continue
            else:
                lst.append( to_system(ws,r,i+1) )
 
            dump_to_json(directory,s,lst)
 
    wb.save( xl_path(directory) )   # There may have been UUIDs added

    # ---------------------------------------------------------------------------
    # Process references 
    # 
    # obj = {
        # '__entry__' : 'Reference',        
        # 'uid' : [], 
        # 'locale' : locale, 
        # 'unit-system' : dict(
            # uid = [],
            # dimensions = [],
            # prefix = []
        # ),
        # "UCUM" : dict(
            # code = "",
            # description = "",
        # )
    # }
    
    directory = "references"
    def to_reference(ws,r,i):

        uid = get_uid(directory,ws,r[0].value,i,0+1)

        obj = dict(
            __entry__ = 'Reference',
            uid = uid,
            locale = dict( 
                default = dict(
                    name = str( r[1].value ),
                    symbol = str() if r[2].value is None else str( r[2].value )
                )
            )
        )

        # A unit system might not be associated with a reference
        if r[3].value is not None:
            obj['system'] = dict(
                uid = put_uid(directory,ws,r[3].value,i,3+1),
                dimensions = str( literal_eval( r[4].value ) ),
                prefix = literal_eval( r[5].value )
            )
           
        # Code in UCUM may not be provided
        if r[6].value is not None:  
            obj['UCUM'] = dict(
                code = str( r[6].value ),
                description = str( r[7].value )
            )
        
        return obj
        
    wb = load_workbook( xl_path(directory), data_only=True  )
    for s in wb.sheetnames:
        lst = []
        ws = wb[s]
        for i,r in enumerate(ws.rows):
            if i == 0: 
                continue
            else:
                lst.append( to_reference(ws,r,i+1) )

        dump_to_json(directory,s,lst)
    
    wb.save( xl_path(directory) )   # There may have been UUIDs added

# ---------------------------------------------------------------------------
    # Process scales 
    # 
    # obj = {
        # '__entry__' : 'Scale',        
        # 'uid' : [], 
        # 'reference' : [],
        # 'scale_type' : "" 
        # 'systematic' : 1
    # }
    
    directory = "scales"
    def to_scale(ws,r,i):

        uid = get_uid(directory,ws,r[0].value,i,0+1)

        obj =  dict(
            __entry__ = 'Scale',
            uid = uid,
            reference = put_uid(directory,ws,r[1].value,i,1+1),
            scale_type = str( r[2].value )
        )
 
        # A unit system might not be associated with a reference
        if r[3].value is not None:
            obj['systematic'] = int( r[3].value )
            
        return obj 
 
    wb = load_workbook( xl_path(directory), data_only=True  )
    for s in wb.sheetnames:
        # print(directory,s)
        lst = []
        ws = wb[s]
        for i,r in enumerate(ws.rows):
            if i == 0: 
                continue
            else:
                # print(r,i)
                lst.append( to_scale(ws,r,i+1) )

        dump_to_json(directory,s,lst)
    
    # ---------------------------------------------------------------------------
    # Process scales_for 
    # 
    # obj = {
        # '__entry__' : 'ScalesForAspect',        
        # 'aspect' : [], 
        # 'src' : [], 
        # 'dst' : [], 
        # 'factors' : [], 
    # }
            
    directory = "scales_for"
    def to_scales_for(ws,r,i):
    
        return dict(
            __entry__ = 'ScalesForAspect',
            aspect = put_uid(directory,ws,r[0].value,i,0+1),
            src = put_uid(directory,ws,r[1].value,i,1+1),
            dst = put_uid(directory,ws,r[2].value,i,2+1),
            factors = literal_eval( r[3].value )
        )
        
    wb = load_workbook( xl_path(directory), data_only=True  )
    for s in wb.sheetnames:
        lst = []
        ws = wb[s]
        for i,r in enumerate(ws.rows):
            if i == 0: 
                continue
            else:
                lst.append( to_scales_for(ws,r,i+1) )
            
        dump_to_json(directory,s,lst)

    wb.save( xl_path(directory) )   # There may have been UUIDs added

    # ---------------------------------------------------------------------------
    # Process conversion 
    # 
    # obj = {
        # '__entry__' : 'Conversion',        
        # 'src' : [], 
        # 'dst' : [], 
        # 'factors' : [], 
    # }
    
    directory = "conversion"
    def to_conversion(ws,r,i):
        return dict(
            __entry__ = 'Conversion',
            src = put_uid(directory,ws,r[0].value,i,0+1),
            dst = put_uid(directory,ws,r[1].value,i,1+1),
            factors = literal_eval( r[2].value )
        )
        
    wb = load_workbook( xl_path(directory), data_only=True  )
    for s in wb.sheetnames:
        lst = []
        ws = wb[s]
        for i,r in enumerate(ws.rows):
            if i == 0: 
                continue
            else:
                lst.append( to_conversion(ws,r,i+1) )
            
        dump_to_json(directory,s,lst)
            
    wb.save( xl_path(directory) )   # There may have been UUIDs added

    # ---------------------------------------------------------------------------
    # Process casting 
    # 
    # obj = {
        # '__entry__' : 'Cast',        
        # 'src' : [[scale,aspect]], 
        # 'dst' : [[scale,aspect]], 
        # 'function' : "",
        # 'parameters' : dict()
    # }
    
    directory = "casting"
    def to_casting(ws,r,i):
        return dict(
            __entry__ = 'Cast',
            src = [
                put_uid(directory,ws,r[0].value,i,0+1),
                put_uid(directory,ws,r[1].value,i,1+1)
            ],
            dst = [
                put_uid(directory,ws,r[2].value,i,2+1),
                put_uid(directory,ws,r[3].value,i,3+1)
            ],
            function = str( r[4].value ),
            parameters = literal_eval( r[5].value )
        )
        
    wb = load_workbook( xl_path(directory), data_only=True  )
    for s in wb.sheetnames:
        lst = []
        ws = wb[s]
        for i,r in enumerate(ws.rows):
            if i == 0: continue
            
            lst.append( to_casting(ws,r,i+1) )   
            
        dump_to_json(directory,s,lst)
 
    wb.save( xl_path(directory) )   # There may have been UUIDs added
 
