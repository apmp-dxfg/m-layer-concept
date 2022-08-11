"""
Convert an existing collection of M-layer register JSON files 
into a set of Excel workbooks.

The intention is to modify these workbooks and then generate 
a new suite of JSON files.

References will be resolved when the sheets are opened in Excel. 
The internal representation of those paths is then changed and 
cannot be seen again.

So after running this script, the set of workbooks must be opened in Excel 
and then saved (at which point the cell contents will be resolved). 

"""
import json 
import glob
import os

from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter 
 
# --------------------------------------------------------------------------- 
root = os.path.abspath( os.path.dirname(__file__) )

json_root = os.path.join( root, r'../json' )
xl_root = os.path.join( root, 'xl' )

# --------------------------------------------------------------------------- 
def ws_label(ws,labels):
    for i,l_i in enumerate(labels):
        ws.cell(row=1,column=i+1,value=l_i)
    
# --------------------------------------------------------------------------- 
def add_uid(file,ws,uid,row,col):

    if tuple(uid) in UID:
        raise RuntimeError(f"{uid} is assigned")
      
    col_letter = get_column_letter(col)
    addr = f"{col_letter}{row}"
    
    UID[tuple(uid)] = f"'{root}\\xl\\[{file}.xlsx]{ws.title}'!{addr}"
    
    ws[addr] = str(uid)
 
# --------------------------------------------------------------------------- 
def add_string(ws,s,row,col):
    ws.cell(row=row,column=col,value=str(s))

# --------------------------------------------------------------------------- 
def add_uid_reference(ws,uid,row,col):
    ref = UID[ tuple(uid) ]
    ws.cell(row=row,column=col,value=f"={ref}")
    
# ---------------------------------------------------------------------------
workbooks = dict() 
def get_wb(directory):

    if directory not in workbooks:
        wb = Workbook()
        wb.remove( wb.active )  # Remove default 'Sheet'
        workbooks[directory] = wb
    
    return workbooks[directory]
        
# =========================================================================== 
if __name__ == '__main__': 

    locale = dict(
        default = dict(
            name="",
            symbol=""
        )
    )
    
    UID = dict()    # Mapping of uid -> sheet address
    
    if not os.path.isdir( xl_root ): os.mkdir( xl_root )

    # ---------------------------------------------------------------------------
    # Process aspects 
    # 
    directory = "aspects"
    obj = {
        '__entry__' : 'Aspect',        
        'uid' : [], 
        'locale' : locale, 
        'metadata' : dict()
    }
    labels = [
        'uid',
        'locale.default.name','locale.default.symbol',
        'metadata.url'
    ]
        
    path = os.path.abspath(
        os.path.join( 
            json_root,
            directory,
            r'*.json'
        )
    )
    
    wb = get_wb(directory)
    for f_json in glob.glob( path ):
        name = os.path.splitext( os.path.basename(f_json) )[0]
        ws = wb.create_sheet()
        ws.title = f'{name}'       
        ws_label(ws,labels)

        try:        
            with open(f_json,'r') as f:
                data = json.load(f)
        except json.decoder.JSONDecodeError as e:
            # Report errors but do not stop execution
            print("json.decoder.JSONDecodeError",e, 'in:',f_json)
            raise          
        
        for i,d_i in enumerate( data ):
            add_uid(directory,ws,d_i['uid'],row=i+2,col=1)
            add_string(ws,d_i['locale']['default']['name'],row=i+2,col=2)
            add_string(ws,d_i['locale']['default']['symbol'],row=i+2,col=3)
            try:
                add_string(ws,d_i['metadata']['url'],row=i+2,col=4)
            except KeyError as k:
                ws.cell(row=i+2,column=4,value=None)
            
    # ---------------------------------------------------------------------------
    # Process systems 
    # 
    directory = "system"
    # obj = {
        # '__entry__' : 'UnitSystem',        
        # 'uid' : [], 
        # 'name' : "", 
        # 'basis' : []
    # }
    
    labels = [
        'uid',
        'name',
        'basis'
    ]
        
    path = os.path.abspath(
        os.path.join( 
            json_root,
            directory,
            r'*.json'
        )
    )

    wb = get_wb(directory)
    for f_json in glob.glob( path ):
        name = os.path.splitext( os.path.basename(f_json) )[0]
        ws = wb.create_sheet()
        ws.title = f'{name}'       
        ws_label(ws,labels)

        try:        
            with open(f_json,'r') as f:
                data = json.load(f)
        except json.decoder.JSONDecodeError as e:
            # Report errors but do not stop execution
            print("json.decoder.JSONDecodeError",e, 'in:',f_json)
            raise          
        
        for i,d_i in enumerate( data ):
            add_uid(directory,ws,d_i['uid'],row=i+2,col=1)
            add_string(ws,d_i['name'],row=i+2,col=2)
            add_string(ws,d_i['basis'],row=i+2,col=3)
            
    # ---------------------------------------------------------------------------
    # Process references 
    # 
    directory = "references"
    # obj = {
        # '__entry__' : 'Reference',        
        # 'uid' : [], 
        # 'locale' : locale, 
        # 'system' : dict(
            # uid = [],
            # dimensions = [],
            # prefix = 1
        # ),
        # "UCUM" : dict(
            # code = "",
            # description = "",
        # )
    # }
    
    labels = [
        'uid',
        'locale.default.name','locale.default.symbol',
        'system.uid', 'system.dimensions', 'system.prefix',
        'UCUM.code','UCUM.description'
    ]
        
    path = os.path.abspath(
        os.path.join( 
            json_root,
            directory,
            r'*.json'
        )
    )

    wb = get_wb(directory)
    for f_json in glob.glob( path ):
        name = os.path.splitext( os.path.basename(f_json) )[0]
        ws = wb.create_sheet()
        ws.title = f'{name}'       
        ws_label(ws,labels)

        try:        
            with open(f_json,'r') as f:
                data = json.load(f)
        except json.decoder.JSONDecodeError as e:
            # Report errors but do not stop execution
            print("json.decoder.JSONDecodeError",e, 'in:',f_json)
            raise         
        
        for i,d_i in enumerate( data ):
            add_uid(directory,ws,d_i['uid'],row=i+2,col=1)
            add_string(ws,d_i['locale']['default']['name'],row=i+2,col=2)
            add_string(ws,d_i['locale']['default']['symbol'],row=i+2,col=3)
            try:    # UnitSystem may not be included
                add_uid_reference(ws,d_i['system']['uid'],row=i+2,col=4)
                add_string(ws,d_i['system']['dimensions'],row=i+2,col=5)
                add_string(ws,d_i['system']['prefix'],row=i+2,col=6)
            except KeyError as k:   
                ws.cell(row=i+2,column=4,value=None)
                ws.cell(row=i+2,column=5,value=None)
                ws.cell(row=i+2,column=6,value=None)
            try:    # UCUM code may not be included
                add_string(ws,d_i['UCUM']['code'],row=i+2,col=7)
                add_string(ws,d_i['UCUM']['description'],row=i+2,col=8)
            except KeyError as k:
                ws.cell(row=i+2,column=7,value=None)
                ws.cell(row=i+2,column=8,value=None)
            
    # ---------------------------------------------------------------------------
    # Process scales 
    # 
    directory = "scales"
    # obj = {
        # '__entry__' : 'Scale',        
        # 'uid' : [], 
        # 'reference' : [],
        # 'scale_type' : "" 
    # }
    
    labels = [
        'uid',
        'reference',
        'scale_type'
    ]
        
    path = os.path.abspath(
        os.path.join( 
            json_root,
            directory,
            r'*.json'
        )
    )

    wb = get_wb(directory)
    for f_json in glob.glob( path ):
        name = os.path.splitext( os.path.basename(f_json) )[0]
        ws = wb.create_sheet()
        ws.title = f'{name}'       
        ws_label(ws,labels)

        try:        
            with open(f_json,'r') as f:
                data = json.load(f)
        except json.decoder.JSONDecodeError as e:
            # Report errors but do not stop execution
            print("json.decoder.JSONDecodeError",e, 'in:',f_json)
            raise 
               
        for i,d_i in enumerate( data ):
            add_uid(directory,ws,d_i['uid'],row=i+2,col=1)
            add_uid_reference(ws,d_i['reference'],row=i+2,col=2)
            add_string(ws,d_i['scale_type'],row=i+2,col=3)
            
    # ---------------------------------------------------------------------------
    # Process scales_for 
    # 
    directory = "scales_for"
    # obj = {
        # '__entry__' : 'ScalesForAspect',        
        # 'aspect' : [], 
        # 'src' : [], 
        # 'dst' : [], 
        # 'factors' : [], 
    # }
    
    labels = [
        'aspect',
        'src',
        'dst',
        'factors'
    ]
        
    path = os.path.abspath(
        os.path.join( 
            json_root,
            directory,
            r'*.json'
        )
    )

    wb = get_wb(directory)
    for f_json in glob.glob( path ):
        name = os.path.splitext( os.path.basename(f_json) )[0]
        ws = wb.create_sheet()
        ws.title = f'{name}'        
        ws_label(ws,labels)

        try:        
            with open(f_json,'r') as f:
                data = json.load(f)
        except json.decoder.JSONDecodeError as e:
            # Report errors but do not stop execution
            print("json.decoder.JSONDecodeError",e, 'in:',f_json)
            raise 
               
        for i,d_i in enumerate( data ):
            add_uid_reference(ws,d_i['aspect'],row=i+2,col=1)
            add_uid_reference(ws,d_i['src'],row=i+2,col=2)
            add_uid_reference(ws,d_i['dst'],row=i+2,col=3)
            add_string(ws,d_i['factors'],row=i+2,col=4)
 
    # ---------------------------------------------------------------------------
    # Process conversion 
    # 
    directory = "conversion"
    # obj = {
        # '__entry__' : 'Conversion',        
        # 'src' : [], 
        # 'dst' : [], 
        # 'factors' : [], 
    # }
    
    labels = [
        'src',
        'dst',
        'factors'
    ]
        
    path = os.path.abspath(
        os.path.join( 
            json_root,
            directory,
            r'*.json'
        )
    )

    wb = get_wb(directory)
    for f_json in glob.glob( path ):
        name = os.path.splitext( os.path.basename(f_json) )[0]
        ws = wb.create_sheet()
        ws.title = f'{name}'        
        ws_label(ws,labels)

        try:        
            with open(f_json,'r') as f:
                data = json.load(f)
        except json.decoder.JSONDecodeError as e:
            # Report errors but do not stop execution
            print("json.decoder.JSONDecodeError",e, 'in:',f_json)
            raise         
        
        for i,d_i in enumerate( data ):
            add_uid_reference(ws,d_i['src'],row=i+2,col=1)
            add_uid_reference(ws,d_i['dst'],row=i+2,col=2)
            add_string(ws,d_i['factors'],row=i+2,col=3) 
            
    # ---------------------------------------------------------------------------
    # Process casting 
    # 
    directory = "casting"
    # obj = {
        # '__entry__' : 'Cast',        
        # 'src' : [], 
        # 'dst' : [], 
        # 'function' : "",
        # 'parameters' : dict()
    # }
    
    labels = [
        'src.scale','src.aspect',
        'dst.scale','dst.aspect',
        'function',
        'parameters'
    ]
        
    path = os.path.abspath(
        os.path.join( 
            json_root,
            directory,
            r'*.json'
        )
    )

    wb = get_wb(directory)
    for f_json in glob.glob( path ):
        name = os.path.splitext( os.path.basename(f_json) )[0]
        ws = wb.create_sheet()
        ws.title = f"{name}"
        ws_label(ws,labels)

        try:        
            with open(f_json,'r') as f:
                data = json.load(f)
        except json.decoder.JSONDecodeError as e:
            # Report errors but do not stop execution
            print("json.decoder.JSONDecodeError",e, 'in:',f_json)
            raise         
        
        for i,d_i in enumerate( data ):
            add_uid_reference(ws,d_i['src'][0],row=i+2,col=1)
            add_uid_reference(ws,d_i['src'][1],row=i+2,col=2)
            add_uid_reference(ws,d_i['dst'][0],row=i+2,col=3)
            add_uid_reference(ws,d_i['dst'][1],row=i+2,col=4)
            add_string(ws,d_i['function'],row=i+2,col=5)             
            add_string(ws,d_i['parameters'],row=i+2,col=6)    
            
    # ---------------------------------------------------------------------------
    # Save workbooks together 
    # 
    root = os.path.abspath( os.path.dirname(__file__) )
    for name_i in workbooks:
        destination = os.path.join( f"{root}", "xl" )
        document_name = f'{name_i}.xlsx'
        filename = os.path.join(destination, document_name)
        workbooks[name_i].save(filename=filename)
    
